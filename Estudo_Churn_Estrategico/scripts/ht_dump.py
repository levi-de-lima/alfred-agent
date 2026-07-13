# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")
F=r"C:\Users\TMB1.NOTE-TMB46\OneDrive - TMB Educação\TMB - Documentos\Comercial\5 - Rev Ops\05_Projetos\AI_Churn\Estudo_Churn_Estrategico\Inadimplencia_HighTicket_TMB_v11.xlsx"
wb=openpyxl.load_workbook(F, read_only=True, data_only=True)
def num(x):
    if isinstance(x,float):
        return (f"{x:.4f}" if abs(x)<10 else f"{x:,.0f}")
    return str(x) if x is not None else "·"
LOGIC=['Parametros','🎓 Por Score','⚠️ Por Risco','📊 Comparativo Métodos','🔮 Projeção','🎯 Por Tipo Produto','📉 Por Parcela','🚻 Por Gênero','Listas','🔍 Guia de Auditoria','📋 Resumo Executivo']
for name in LOGIC:
    if name not in wb.sheetnames:
        print("!! falta",name); continue
    ws=wb[name]
    print("\n########", name, "(",ws.max_row,"x",ws.max_column,") ########")
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i>=45: print("   ...(trunc)"); break
        cells=[num(c) for c in row[:ws.max_column]]
        if any(c!="·" for c in cells):
            print(" | ".join(cells))
wb.close()
